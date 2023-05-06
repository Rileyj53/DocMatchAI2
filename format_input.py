from openpyxl import load_workbook
import spacy
import json


# Define a function to extract words and their positions from text
def extract_words(nlp, text, word):
    # print(word)
    # check if the word with an 's' is present in the text
    if f"{word}s" in text.lower():
        word = f"{word}s"

    doc = nlp(text)

    print(word)
    print(text)
    if ' ' in word:
        new_word_list = []
        # Split the word into two separate words
        new_words = word.split()
        # Append the two new words to the new list
        new_word_list.extend(new_words)
        word_start, word_end = new_word_list[0], new_word_list[-1]
        for token in doc:
            if token.text.lower() == word_start:
                start_pos = token.idx
            if token.text.lower() == word_end:
                end_pos = token.idx + len(token)
                car_pos = start_pos, end_pos

    else:
        for token in doc:
            if token.text.lower() == word.lower():
                car_pos = token.idx, token.idx + len(token)

    try:
        return car_pos
    except UnboundLocalError:
        print(word)
        print(text)
        exit(1)


# Define the main function for processing the input Excel file and generating the output JSON file
def input_and_convert(path, input_xlsx, output_txt):
    # Open Excel file and convert it
    workbook = load_workbook(path + input_xlsx)

    # Load the spacy language model for English
    nlp = spacy.load('en_core_web_sm')

    with open(path + output_txt, "w") as f:
        # Iterate over each worksheet in the Excel file
        for sheetname in workbook.sheetnames:
            worksheet = workbook[sheetname]

            # Iterate over each row in the worksheet
            for row in worksheet.iter_rows():
                # Convert words in the first column to lowercase and remove leading whitespace in column B
                updated_row = [
                    cell.value.strip() if cell.column == 2 else cell.value.lower() if cell.column == 1 else cell.value
                    for cell in row]
                text = updated_row[1].strip()
                word_list = updated_row[0].strip().lower().split(', ')

                entity_list = []
                for word in word_list:
                    car_pos = extract_words(nlp, text, word)
                    type_of = 'MEDICAL_PROFESSION'
                    entity_list.append((car_pos, type_of))

                # Generate the output dictionary in the required format
                output_dict = {"entities": [(pos[0], pos[1], label) for pos, label in entity_list]}
                output_json = json.dumps(output_dict)

                # Write the output to the output file
                output = f'"{text}", {output_json}\n'
                f.write(output)
